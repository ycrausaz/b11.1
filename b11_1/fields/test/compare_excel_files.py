# compare_excel_files.py

import pandas as pd
import csv
import argparse
from openpyxl import load_workbook

def read_mapping_file(mapping_file):
    mapping = {}
    with open(mapping_file, 'r') as file:
        reader = csv.reader(file)
        for row in reader:
            file1_sheet = row[0].strip()
            file2_sheet = row[1].strip()
            start_row = int(row[2].strip())
            header_row = int(row[3].strip())
            mapping[file1_sheet] = (file2_sheet, start_row, header_row)
    return mapping

def find_column_index_by_header(sheet, header_row, header_name):
    for col in range(1, sheet.max_column + 1):
        if sheet.cell(row=header_row, column=col).value == header_name:
            return col
    return None

def normalize_value(value):
    if value is None or (isinstance(value, str) and value.strip().lower() in ('null', 'none', '')) or (isinstance(value, (int, float)) and value == 0):
        return 0.0
    try:
        return float(value)
    except ValueError:
        return value

def compare_sheets(sheet1, sheet2, start_row, header_row):
    headers1 = [sheet1.cell(row=header_row, column=col).value for col in range(1, sheet1.max_column + 1)]
    headers2 = [sheet2.cell(row=header_row, column=col).value for col in range(1, sheet2.max_column + 1)]
    
    common_headers = set(headers1).intersection(headers2)
    col_indices1 = {header: find_column_index_by_header(sheet1, header_row, header) for header in common_headers}
    col_indices2 = {header: find_column_index_by_header(sheet2, header_row, header) for header in common_headers}
    
    differences = []
    
    for row in range(start_row, max(sheet1.max_row, sheet2.max_row) + 1):
        for header in common_headers:
            col1 = col_indices1[header]
            col2 = col_indices2[header]
            cell1 = normalize_value(sheet1.cell(row=row, column=col1).value)
            cell2 = normalize_value(sheet2.cell(row=row, column=col2).value)
            if cell1 != cell2:
                differences.append((row, header, cell1, cell2))
    
    return differences

def compare_excel_files(input_file_1, input_file_2, mapping_file):
    mapping = read_mapping_file(mapping_file)
    
    wb1 = load_workbook(input_file_1)
    wb2 = load_workbook(input_file_2)
    
    all_differences = {}
    
    for sheet1_name, (sheet2_name, start_row, header_row) in mapping.items():
        if sheet1_name in wb1.sheetnames and sheet2_name in wb2.sheetnames:
            sheet1 = wb1[sheet1_name]
            sheet2 = wb2[sheet2_name]
            
            differences = compare_sheets(sheet1, sheet2, start_row, header_row)
            if differences:
                all_differences[(sheet1_name, sheet2_name)] = differences
    
    return all_differences

def main():
    parser = argparse.ArgumentParser(description="Compare two Excel files sheet by sheet and cell by cell using headers for column lookup.")
    parser.add_argument("input_file_1", help="The first input Excel file.")
    parser.add_argument("input_file_2", help="The second input Excel file.")
    parser.add_argument("mapping_file", help="The CSV file that maps sheets, start rows, and header rows for comparison.")
    
    args = parser.parse_args()
    
    differences = compare_excel_files(args.input_file_1, args.input_file_2, args.mapping_file)
    
    if differences:
        for sheets, diff in differences.items():
            print(f"Differences between sheets: {sheets[0]} and {sheets[1]}")
            for d in diff:
                print(f"Row: {d[0]}, Header: {d[1]}, {sheets[0]}: {d[2]}, {sheets[1]}: {d[3]}")
    else:
        print("No differences found.")

if __name__ == "__main__":
    main()

