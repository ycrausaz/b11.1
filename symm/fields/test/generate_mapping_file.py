# generate_mapping_file.py

import pandas as pd
import argparse
import csv
from openpyxl import load_workbook

def generate_mapping_file(input_file_1, input_file_2, output_mapping_file):
    wb1 = load_workbook(input_file_1, read_only=True)
    wb2 = load_workbook(input_file_2, read_only=True)

    sheets1 = wb1.sheetnames
    sheets2 = wb2.sheetnames

    mapping = []

    for sheet in sheets1:
        if sheet in sheets2:
            mapping.append((sheet, sheet))

    with open(output_mapping_file, 'w', newline='') as file:
        writer = csv.writer(file)
        for sheet1, sheet2 in mapping:
            writer.writerow([sheet1, sheet2])

    print(f"Mapping file '{output_mapping_file}' generated successfully.")

def main():
    parser = argparse.ArgumentParser(description="Generate a mapping file for sheet names between two Excel files.")
    parser.add_argument("input_file_1", help="The first input Excel file.")
    parser.add_argument("input_file_2", help="The second input Excel file.")
    parser.add_argument("output_mapping_file", help="The output CSV file for the mapping.")
    
    args = parser.parse_args()
    
    generate_mapping_file(args.input_file_1, args.input_file_2, args.output_mapping_file)

if __name__ == "__main__":
    main()

